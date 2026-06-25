import openpyxl as xl
wb = xl.load_workbook('file_xs1.xlsx')
sheet = wb['Arkusz1']
cell = sheet['a1'] 
cell = sheet.cell(1,1)
#print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = float(cell.value) * 0.9
    corrected_line_price = sheet.cell(row, 4)
    corrected_line_price.value = corrected_price

wb.save('transaction2.xlsx')